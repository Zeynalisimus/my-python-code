from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Создание нового файла Excel
wb = Workbook()
ws = wb.active
ws.title = "Фонд оплаты труда"

# Данные для таблицы
data = [
    ("Сотрудник", "Должность", "Оклад (в месяц)", "Премии (в месяц)"),
    ("Иван Иванов", "Менеджер", 50000, 5000),
    ("Мария Сидорова", "Бухгалтер", 45000, 4000),
    ("Петр Петров", "Продавец", 35000, 3000),
    ("Анна Иванова", "Секретарь", 30000, 2000)
]

# Заполнение таблицы данными
for row_idx, row_data in enumerate(data):
    for col_idx, cell_data in enumerate(row_data):
        cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_data)

# Добавление формул в ячейку по колонке заработной платы
for row in range(2, len(data) + 2):
    ws[f"E{row}"] = f"=C{row}+D{row}"

# Сохранение файла Excel
wb.save("фонд_оплаты_труда.xlsx")
